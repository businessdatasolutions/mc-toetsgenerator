import io

from openpyxl import load_workbook

from .schemas import ParsedOption, ParsedQuestion

OPTION_COLUMNS = ["optie_a", "optie_b", "optie_c", "optie_d"]
OPTION_LABELS = ["A", "B", "C", "D"]
REQUIRED_COLUMNS = {"stam", "optie_a", "optie_b", "optie_c", "optie_d", "correct"}
OPTIONAL_COLUMNS = {
    "categorie": "category",
    "bloom_niveau": "bloom_level",
    "leerdoel": "learning_goal",
    "vraag_id": "question_id",
    "id": "question_id",
}


def parse_xlsx(content: bytes) -> list[ParsedQuestion]:
    """Parse an Excel file with columns: stam, optie_a, optie_b, optie_c, optie_d, correct."""
    wb = load_workbook(filename=io.BytesIO(content), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel bestand is leeg")

    # Normalize headers
    headers = [str(cell).strip().lower() if cell else "" for cell in rows[0]]
    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        raise ValueError(
            f"Ontbrekende kolommen in Excel: {', '.join(sorted(missing))}. "
            f"Verwacht: {', '.join(sorted(REQUIRED_COLUMNS))}"
        )

    col_idx = {name: i for i, name in enumerate(headers)}

    questions: list[ParsedQuestion] = []
    for row_num, row in enumerate(rows[1:], start=2):
        stem = str(row[col_idx["stam"]] or "").strip()
        if not stem:
            continue

        correct_label = str(row[col_idx["correct"]] or "").strip().upper()
        if correct_label not in OPTION_LABELS:
            raise ValueError(
                f"Rij {row_num}: ongeldige waarde voor 'correct': '{correct_label}'. "
                f"Verwacht: A, B, C of D"
            )

        options: list[ParsedOption] = []
        for i, col_name in enumerate(OPTION_COLUMNS):
            text = str(row[col_idx[col_name]] or "").strip()
            options.append(
                ParsedOption(
                    text=text,
                    position=i,
                    is_correct=(OPTION_LABELS[i] == correct_label),
                )
            )

        # Read optional columns
        extra: dict[str, str | None] = {}
        for xlsx_col, field_name in OPTIONAL_COLUMNS.items():
            if xlsx_col in col_idx and field_name not in extra:
                val = str(row[col_idx[xlsx_col]] or "").strip()
                if val:
                    extra[field_name] = val

        # Auto-assign question_id from row number if not in file
        if "question_id" not in extra:
            extra["question_id"] = str(row_num - 1)

        questions.append(ParsedQuestion(stem=stem, options=options, **extra))

    wb.close()
    return questions
